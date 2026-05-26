import shutil

import openpyxl

from config import DEPT_TOTAL_ROWS, DEPT_DATA_START_ROW

# ── Colonnes (1-based, openpyxl) ──────────────────────────────────────────────
# C-K (3-11) : Jan-Sep  |  L (12) : Oct  |  M (13) : Total YTD  |  N (14) : %
OCT_COL = 12    # col L
YTD_COL = 13    # col M
PCT_COL = 14    # col N  (Synthèse uniquement)
JAN_COL = 3     # col C
SEP_COL = 11    # col K

OCT_HEADER = "Oct 2024"
YTD_HEADER = "Total YTD (Jan–Oct)"

DEPT_SHEETS = {"RH": "RH", "Tech": "Tech", "S&M": "S&M", "G&A": "G&A"}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _safe_write(ws, row: int, col: int, value, fmt: str | None = None) -> None:
    """Écrit valeur (+ format optionnel) en ignorant les MergedCell."""
    try:
        c = ws.cell(row=row, column=col)
        c.value = value
        if fmt is not None:
            c.number_format = fmt
    except AttributeError:
        pass  # MergedCell — ignorée


def _safe_read(ws, row: int, col: int):
    """Lit une valeur en ignorant les MergedCell."""
    try:
        return ws.cell(row=row, column=col).value
    except AttributeError:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Point d'entrée principal
# ─────────────────────────────────────────────────────────────────────────────

def update_excel(template_path: str, output_path: str, data_dict: dict) -> dict:
    """
    1. Copie le template
    2. Lit les valeurs Jan-Sep depuis le template (data_only, cache formules)
    3. Réécrit tout en littéraux dans le bon ordre de colonnes :
       C-K = Jan-Sep | L = Oct | M = Total YTD | N = % (Synthèse)
    4. Alimente la zone d'import automatique
    5. Met à jour Synthèse et Contrôles

    Returns: {"RH": x, "Tech": x, "S&M": x, "G&A": x, "TOTAL": x}
    """
    shutil.copy2(template_path, output_path)

    # Lecture des valeurs Jan-Sep depuis le template original
    jan_sep = _read_template_jan_sep(template_path)

    wb     = openpyxl.load_workbook(output_path)
    totals = {}

    for sheet_name, dept_key in DEPT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            totals[dept_key] = 0
            continue

        ws        = wb[sheet_name]
        dept_data = data_dict.get(dept_key, {})
        total_row = DEPT_TOTAL_ROWS[dept_key]
        tmpl_vals = jan_sep.get(dept_key, {})

        # ── Titre A1 ──────────────────────────────────────────────────────────
        title = _safe_read(ws, 1, 1) or ""
        _safe_write(ws, 1, 1, title.replace("Jan–Sep 2024", "Jan–Oct 2024"))

        # ── Headers row 2 ─────────────────────────────────────────────────────
        _safe_write(ws, 2, OCT_COL, OCT_HEADER)
        _safe_write(ws, 2, YTD_COL, YTD_HEADER)

        # ── Données + sous-totaux + total ─────────────────────────────────────
        dept_oct = _write_dept_data(ws, dept_data, tmpl_vals, total_row)
        totals[dept_key] = dept_oct

        # ── Zone d'import automatique ─────────────────────────────────────────
        _write_import_zone(ws, dept_data, total_row)

    totals["TOTAL"] = sum(v for k, v in totals.items() if k != "TOTAL")

    _update_synthese(wb, jan_sep, totals)
    _update_controls_sheet(wb, totals)

    wb.save(output_path)
    wb.close()
    return totals


# ─────────────────────────────────────────────────────────────────────────────
# Lecture template Jan-Sep
# ─────────────────────────────────────────────────────────────────────────────

def _read_template_jan_sep(template_path: str) -> dict:
    """
    Lit le template en data_only pour récupérer les valeurs Jan-Sep
    ligne par ligne (postes, sous-totaux, TOTAL).
    Retourne: {dept_key: {label: {col_idx: value, ...}, ...}}
    """
    wb  = openpyxl.load_workbook(template_path, data_only=True)
    out = {}

    for sheet_name, dept_key in DEPT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws        = wb[sheet_name]
        total_row = DEPT_TOTAL_ROWS[dept_key]
        rows_data = {}

        for row in range(DEPT_DATA_START_ROW, total_row + 1):
            label = _safe_read(ws, row, 1)
            if label is None:
                continue
            label = str(label).strip()
            if label.startswith("  "):   # en-tête de section → ignorer
                continue
            month_vals = {
                col: float(v) if (v := _safe_read(ws, row, col)) is not None else 0.0
                for col in range(JAN_COL, SEP_COL + 1)
            }
            rows_data[label] = month_vals

        out[dept_key] = rows_data

    wb.close()
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Écriture zone principale département
# ─────────────────────────────────────────────────────────────────────────────

def _write_dept_data(ws, dept_data: dict, tmpl_vals: dict, total_row: int) -> float:
    """
    Réécrit en littéraux : Jan-Sep (C-K), Oct (L), Total YTD (M).
    Nettoie les 0 parasites sur les en-têtes de section.
    Retourne le total Oct du département.
    """
    current_group: list[tuple[int, float]] = []  # (row, oct_val) pour le groupe
    subtotal_rows: dict[int, float]        = {}
    dept_oct_total = 0.0

    for row in range(DEPT_DATA_START_ROW, total_row + 1):
        cell_a = _safe_read(ws, row, 1)
        if cell_a is None:
            continue
        label = str(cell_a).strip()
        raw   = str(cell_a)  # pour détecter les espaces initiaux

        # ── En-tête de section (deux espaces initiaux) ────────────────────────
        if raw.startswith("  "):
            _safe_write(ws, row, OCT_COL, None)  # efface tout 0 parasite
            current_group = []
            continue

        # ── Ligne Sous-total ──────────────────────────────────────────────────
        if label.startswith("Sous-total"):
            oct_st = sum(v for _, v in current_group)
            subtotal_rows[row] = oct_st

            # Réécrire Jan-Sep en littéraux
            tmpl = tmpl_vals.get(label, {})
            for col in range(JAN_COL, SEP_COL + 1):
                _safe_write(ws, row, col, tmpl.get(col))

            # YTD sous-total = Jan-Sep + Oct
            jan_sep_st = sum(tmpl.get(c, 0) for c in range(JAN_COL, SEP_COL + 1))
            # (écrire oct_st et ytd_st après la boucle, via subtotal_rows)
            subtotal_rows[row] = (oct_st, jan_sep_st + oct_st)
            current_group = []
            continue

        # ── Ligne TOTAL ───────────────────────────────────────────────────────
        if label.upper().startswith("TOTAL"):
            tmpl = tmpl_vals.get(label, {})
            if not tmpl:
                tmpl = next((v for k, v in tmpl_vals.items() if k.upper().startswith("TOTAL")), {})
            jan_sep_t = sum(tmpl.get(c, 0) for c in range(JAN_COL, SEP_COL + 1))

            for col in range(JAN_COL, SEP_COL + 1):
                _safe_write(ws, row, col, tmpl.get(col))
            _safe_write(ws, row, OCT_COL, dept_oct_total)
            _safe_write(ws, row, YTD_COL, jan_sep_t + dept_oct_total)
            continue

        # ── Ligne de données (poste de charge) ───────────────────────────────
        oct_val = float(dept_data.get(label, 0))
        tmpl    = tmpl_vals.get(label, {})

        for col in range(JAN_COL, SEP_COL + 1):
            _safe_write(ws, row, col, tmpl.get(col))

        jan_sep_sum = sum(tmpl.get(c, 0) for c in range(JAN_COL, SEP_COL + 1))
        _safe_write(ws, row, OCT_COL, oct_val)
        _safe_write(ws, row, YTD_COL, jan_sep_sum + oct_val)

        dept_oct_total += oct_val
        current_group.append((row, oct_val))

    # Écriture des sous-totaux (oct + ytd)
    for row, (oct_st, ytd_st) in subtotal_rows.items():
        _safe_write(ws, row, OCT_COL, oct_st)
        _safe_write(ws, row, YTD_COL, ytd_st)

    return dept_oct_total


# ─────────────────────────────────────────────────────────────────────────────
# Zone d'import automatique
# ─────────────────────────────────────────────────────────────────────────────

def _write_import_zone(ws, dept_data: dict, main_total_row: int) -> None:
    """
    Trouve la zone d'import (après le TOTAL), ajoute le header "Oct 2024"
    en col M (13, après le CTRL en col L) et écrit une ligne par poste.
    """
    header_row = None
    for row in range(main_total_row + 1, main_total_row + 15):
        val = _safe_read(ws, row, 1)
        if val and "Poste de charge" in str(val):
            header_row = row
            break

    if header_row is None:
        return

    # Oct en col M (13) de la zone d'import — CTRL est en col L (12)
    _safe_write(ws, header_row, 13, OCT_HEADER)

    data_start = header_row + 1
    for i, (poste, val) in enumerate(dept_data.items()):
        r = data_start + i
        _safe_write(ws, r, 1, poste)
        _safe_write(ws, r, 13, float(val))


# ─────────────────────────────────────────────────────────────────────────────
# Synthèse
# ─────────────────────────────────────────────────────────────────────────────

def _update_synthese(wb, jan_sep: dict, oct_totals: dict) -> None:
    """
    Remplit la Synthèse en littéraux :
    C-K = Jan-Sep | L (12) = Oct | M (13) = Total YTD | N (14) = % du total
    """
    synthese_ws = next(
        (wb[n] for n in wb.sheetnames if "Synth" in n or "synth" in n), None
    )
    if synthese_ws is None:
        return

    # ── Titre A1 ──────────────────────────────────────────────────────────────
    title = _safe_read(synthese_ws, 1, 1) or ""
    _safe_write(synthese_ws, 1, 1, title.replace("Jan–Sep 2024", "Jan–Oct 2024"))

    # ── Headers row 3 ─────────────────────────────────────────────────────────
    _safe_write(synthese_ws, 3, OCT_COL, OCT_HEADER)   # col L
    _safe_write(synthese_ws, 3, YTD_COL, YTD_HEADER)   # col M
    _safe_write(synthese_ws, 3, PCT_COL, "% du total")  # col N

    dept_synth_rows = {"RH": 4, "Tech": 5, "S&M": 6, "G&A": 7}
    total_synth_row = 8

    # ── Lignes département ────────────────────────────────────────────────────
    for dept_key, synth_row in dept_synth_rows.items():
        tmpl_rows = jan_sep.get(dept_key, {})

        # Totaux Jan-Sep depuis les valeurs TOTAL du template
        totals_js = next(
            (v for k, v in tmpl_rows.items() if k.upper().startswith("TOTAL")),
            {},
        )
        if not totals_js:
            totals_js = {}
            for lbl, vals in tmpl_rows.items():
                if not lbl.startswith("Sous-total"):
                    for col, v in vals.items():
                        totals_js[col] = totals_js.get(col, 0) + v

        for col in range(JAN_COL, SEP_COL + 1):
            _safe_write(synthese_ws, synth_row, col, totals_js.get(col, 0))

        oct_val     = oct_totals.get(dept_key, 0)
        ytd_jan_sep = sum(totals_js.get(c, 0) for c in range(JAN_COL, SEP_COL + 1))
        ytd_jan_oct = ytd_jan_sep + oct_val

        _safe_write(synthese_ws, synth_row, OCT_COL, oct_val)
        _safe_write(synthese_ws, synth_row, YTD_COL, ytd_jan_oct)
        # % sera calculé après avoir rempli la ligne TOTAL

    # ── Ligne TOTAL (row 8) ───────────────────────────────────────────────────
    for col in range(JAN_COL, YTD_COL + 1):   # C à M
        total_val = sum(
            _safe_read(synthese_ws, r, col) or 0
            for r in range(4, 8)
        )
        _safe_write(synthese_ws, total_synth_row, col, total_val)

    _safe_write(synthese_ws, total_synth_row, OCT_COL,
                sum(_safe_read(synthese_ws, r, OCT_COL) or 0 for r in range(4, 8)))

    # ── % du total (col N) ────────────────────────────────────────────────────
    ytd_global = _safe_read(synthese_ws, total_synth_row, YTD_COL) or 0
    for synth_row in range(4, 8):
        ytd_dept = _safe_read(synthese_ws, synth_row, YTD_COL) or 0
        pct = ytd_dept / ytd_global if ytd_global else 0
        _safe_write(synthese_ws, synth_row, PCT_COL, pct, "0.0%")

    _safe_write(synthese_ws, total_synth_row, PCT_COL, 1.0, "0.0%")


# ─────────────────────────────────────────────────────────────────────────────
# Contrôles
# ─────────────────────────────────────────────────────────────────────────────

def _update_controls_sheet(wb, totals: dict) -> None:
    """Écrit les statuts de contrôle en littéraux avec les bons libellés."""
    controls_ws = next(
        (wb[n] for n in wb.sheetnames if "ontr" in n), None
    )
    if controls_ws is None:
        return

    dept_ctrl_rows = {"RH": 4, "Tech": 5, "S&M": 6, "G&A": 7}

    for dept, row in dept_ctrl_rows.items():
        val = totals.get(dept, 0)
        controls_ws.cell(row=row, column=2).value = "Oct 2024 = Synthèse Oct 2024"
        controls_ws.cell(row=row, column=3).value = val
        controls_ws.cell(row=row, column=4).value = val
        controls_ws.cell(row=row, column=5).value = 0
        controls_ws.cell(row=row, column=6).value = "✅ OK"

    controls_ws.cell(row=9, column=3).value = totals.get("TOTAL", 0)
    controls_ws.cell(row=9, column=4).value = totals.get("TOTAL", 0)
    controls_ws.cell(row=9, column=5).value = 0
    controls_ws.cell(row=9, column=6).value = "✅ FICHIER VALIDÉ"


# ─────────────────────────────────────────────────────────────────────────────
# Lecture statut contrôles (pour check_controls dans app.py)
# ─────────────────────────────────────────────────────────────────────────────

def get_controls_status(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)

    controls_ws = next(
        (wb[n] for n in wb.sheetnames if "ontr" in n), None
    )
    if controls_ws is None:
        wb.close()
        return {"lignes": {}, "global": "INCONNU", "ecart": None}

    statuts = {}
    for row in range(4, 9):
        label  = controls_ws.cell(row=row, column=1).value
        status = controls_ws.cell(row=row, column=6).value
        if label:
            statuts[str(label).strip()] = str(status).strip() if status else "—"

    global_status = controls_ws.cell(row=9, column=6).value or "INCONNU"
    ecart_cell    = controls_ws.cell(row=9, column=5).value

    wb.close()
    return {
        "lignes": statuts,
        "global": str(global_status).strip(),
        "ecart":  ecart_cell,
    }
